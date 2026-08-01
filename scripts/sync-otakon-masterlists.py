#!/usr/bin/env python3
"""Fetch Otakon AA + Dealers Google Sheet masterlists → data/ + public/samples/.

Also enriches public/samples/otakon-2026-{artist-alley,dealers}.json with
per-booth catalogInfo used by the shared catalog / Postgres seed.
"""
from __future__ import annotations

import json
import re
import urllib.request
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
AA_SID = "1DVBjS7l8N7LNCQO67wh502dI9iNJdar2s8LwXrsX4OE"
DH_SID = "1MvZBm7FoOcmXYzgV-rUyGDzmQAIKnhF9FlbV0mM7MLs"
AA_URL = f"https://docs.google.com/spreadsheets/d/{AA_SID}/htmlview"
DH_URL = f"https://docs.google.com/spreadsheets/d/{DH_SID}/htmlview"

MEDIA_LABELS = [
    "Anime/Manga",
    "Video Games",
    "Manhua / Manhwa / CJK media",
    "Vtubers / Vocaloid",
    "IRL / Live action",
    "Misc media",
    "OC / other",
]


def clean(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s or None


def booth_keys(raw):
    if raw is None:
        return []
    if isinstance(raw, float) and raw.is_integer():
        return [str(int(raw))]
    s = str(raw).strip()
    if not s:
        return []
    keys = []
    for p in re.split(r"\s*\+\s*", s):
        p = p.strip()
        if not p:
            continue
        if re.fullmatch(r"\d+\.0", p):
            p = str(int(float(p)))
        keys.append(p)
    return keys


def header_has_oc_column(headers: list[str]) -> bool:
    for h in headers:
        hl = h.lower().strip()
        if hl.startswith("oc") or "oc/etc" in hl or hl == "oc/etc merch":
            return True
    return False


def artist_from_vals(vals, *, nsfw_tab=False, has_oc=True):
    """Sheet columns A–L. AA: OC at K, 18+ at L. Dealers: 18+ at K, no OC."""
    if has_oc:
        media_idxs = list(range(4, 11))  # E–K
        labels = MEDIA_LABELS
        adult = clean(vals[11] if len(vals) > 11 else None)
    else:
        media_idxs = list(range(4, 10))  # E–J
        labels = MEDIA_LABELS[:6]
        adult = clean(vals[10] if len(vals) > 10 else None)
        if nsfw_tab and not adult:
            adult = "18+ area (NSFW hall)"

    cats = []
    for i, label in enumerate(labels):
        idx = media_idxs[i]
        val = clean(vals[idx] if idx < len(vals) else None)
        if val:
            cats.append({"label": label, "value": val})

    out = {
        "name": clean(vals[1] if len(vals) > 1 else None),
        "socials": clean(vals[2] if len(vals) > 2 else None),
        "merch": clean(vals[3] if len(vals) > 3 else None),
        "categories": cats or None,
        "adultContent": adult,
    }
    return {k: v for k, v in out.items() if v is not None}


def download(sid: str, dest: Path) -> None:
    url = f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"
    urllib.request.urlretrieve(url, dest)
    print("downloaded", dest, dest.stat().st_size)


def extract_aa(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = [s for s in wb.sheetnames if s.strip().lower().startswith("row")]
    by: dict[str, dict] = {}
    for sheet in sheets:
        ws = wb[sheet]
        current: list[str] | None = None
        for r in range(3, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, 13)]
            if not any(clean(v) for v in vals):
                continue
            keys = booth_keys(vals[0])
            artist = artist_from_vals(vals, has_oc=True)
            if not any(
                artist.get(k) for k in ("name", "merch", "categories", "adultContent", "socials")
            ):
                continue
            if keys:
                for k in keys:
                    entry = {"booth": k, "sheet": sheet, **artist, "tablemates": []}
                    if len(keys) > 1:
                        entry["multiBooth"] = keys
                    by[k] = entry
                current = keys
            elif current:
                mate = dict(artist)
                if mate.get("name"):
                    for k in current:
                        by[k].setdefault("tablemates", []).append(mate)
    for e in by.values():
        if not e.get("tablemates"):
            e.pop("tablemates", None)
    return {
        "source": "unofficial Otakon AA masterlist 2026",
        "sourceUrl": AA_URL,
        "booths": sorted(by.values(), key=lambda x: x["booth"]),
    }


def extract_dh(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    skip = {"start here", "DH map", "InfoCredits"}
    sheets = []
    for s in wb.sheetnames:
        if s in skip:
            continue
        ws = wb[s]
        if str(ws.cell(2, 1).value or "").strip().lower() == "booth":
            sheets.append(s)

    by: dict[str, dict] = {}
    for sheet in sheets:
        ws = wb[sheet]
        nsfw = "nsfw" in sheet.lower()
        headers = [str(ws.cell(2, c).value or "").strip() for c in range(1, 13)]
        has_oc = header_has_oc_column(headers)
        current: list[str] | None = None
        for r in range(3, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, 13)]
            if not any(clean(v) for v in vals):
                continue
            keys = booth_keys(vals[0])
            artist = artist_from_vals(vals, nsfw_tab=nsfw, has_oc=has_oc)
            if not any(
                artist.get(k) for k in ("name", "merch", "categories", "adultContent", "socials")
            ):
                continue
            if keys:
                for k in keys:
                    entry = {"booth": k, "sheet": sheet, **artist, "tablemates": []}
                    if len(keys) > 1:
                        entry["multiBooth"] = keys
                    by[k] = entry
                current = keys
            elif current:
                mate = dict(artist)
                if mate.get("name"):
                    for k in current:
                        by[k].setdefault("tablemates", []).append(mate)
    for e in by.values():
        if not e.get("tablemates"):
            e.pop("tablemates", None)
    return {
        "source": "unofficial Otakon Dealers Hall masterlist 2026",
        "sourceUrl": DH_URL,
        "booths": sorted(by.values(), key=lambda x: (len(x["booth"]), x["booth"])),
    }


def enrich_sample(sample_path: Path, master: dict) -> tuple[int, int]:
    sample = json.loads(sample_path.read_text())
    by = {b["booth"]: b for b in master["booths"]}

    def lookup(booth_id: str):
        if booth_id in by:
            return by[booth_id]
        m = re.match(r"^(\d+)\s*-", booth_id)
        if m and m.group(1) in by:
            return by[m.group(1)]
        return None

    matched = 0
    for b in sample["booths"]:
        info = lookup(b["id"])
        if not info:
            continue
        matched += 1
        if info.get("name"):
            b["name"] = info["name"]
        ci = {
            "source": master["source"],
            "sourceUrl": master["sourceUrl"],
        }
        for key in (
            "socials",
            "merch",
            "categories",
            "adultContent",
            "tablemates",
            "multiBooth",
            "sheet",
        ):
            if info.get(key):
                ci[key] = info[key]
        b["catalogInfo"] = ci
    sample_path.write_text(json.dumps(sample, indent=2, ensure_ascii=False) + "\n")
    return matched, len(sample["booths"])


def main() -> None:
    aa_xlsx = Path("/tmp/otakon-aa-masterlist.xlsx")
    dh_xlsx = Path("/tmp/otakon-dh-masterlist.xlsx")
    download(AA_SID, aa_xlsx)
    download(DH_SID, dh_xlsx)

    aa = extract_aa(aa_xlsx)
    dh = extract_dh(dh_xlsx)

    targets = [
        ROOT / "data/otakon-2026-artist-alley-masterlist.json",
        ROOT / "data/otakon-2026-dealers-masterlist.json",
        ROOT / "public/samples/otakon-2026-artist-alley-masterlist.json",
        ROOT / "public/samples/otakon-2026-dealers-masterlist.json",
    ]
    payloads = [aa, dh, aa, dh]
    for path, payload in zip(targets, payloads):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
        print("wrote", path.relative_to(ROOT), "booths", len(payload["booths"]))

    aa_m, aa_t = enrich_sample(ROOT / "public/samples/otakon-2026-artist-alley.json", aa)
    dh_m, dh_t = enrich_sample(ROOT / "public/samples/otakon-2026-dealers.json", dh)
    for src, dst in [
        (
            ROOT / "public/samples/otakon-2026-artist-alley.json",
            ROOT / "data/otakon-2026-artist-alley-import.json",
        ),
        (
            ROOT / "public/samples/otakon-2026-dealers.json",
            ROOT / "data/otakon-2026-dealers-import.json",
        ),
    ]:
        dst.write_text(src.read_text())
    print(f"enriched AA {aa_m}/{aa_t} DH {dh_m}/{dh_t}")
    print(
        "AA adult/tablemates",
        sum(1 for b in aa["booths"] if b.get("adultContent")),
        sum(1 for b in aa["booths"] if b.get("tablemates")),
    )
    print(
        "DH adult/nsfw-sheet",
        sum(1 for b in dh["booths"] if b.get("adultContent")),
        sum(1 for b in dh["booths"] if "nsfw" in (b.get("sheet") or "").lower()),
    )


if __name__ == "__main__":
    main()
